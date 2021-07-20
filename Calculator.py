# -*- coding: utf-8 -*-
"""
Created on Wed Apr  7 10:28:46 2021

@author: Xylia.Zhang
"""

from tkinter import *
import pandas as pd
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Font,colors,Alignment,PatternFill,Border,Side,Protection


window = Tk()
window.title('Nutrition Facts')

window.geometry('550x750')

tk.Label(window, text='Wish Lulu a nice day  ^__________^',  font=('Arial', 14), width=30, height=2).pack()

tk.Label(window, text="                Formular ID / Formular Name",  font=('Arial', 12)).pack()
title = Frame(window)
title.pack(side=TOP)
tk.Label(title, text="        ",  font=('Arial', 12)).pack(side=LEFT)
f_id = tk.Entry(title, show=None, font=('Arial', 10), width=20)
f_id.pack(side=LEFT)
f_name = tk.Entry(title, show=None, font=('Arial', 10), width=20)
f_name.pack(side=RIGHT) 
lab_item = Label(window, width=100,  height=1,text="                           ID                                          g / pcs         ")
lab_item.pack()

def makeform(window):
    entries = []
    for field in range(1,26):
        row = Frame(window)
        lab = Label(row, width=5, text=field)
        ent = Entry(row, font=('Arial', 10))
        ent_2 = Entry(row, font=('Arial', 10))
        row.pack(side=TOP)
        lab.pack(side=LEFT)
        ent.pack(side=LEFT, fill=X)
        ent_2.pack(side=RIGHT, fill=X)
        entries.append((field, ent, ent_2))

    return entries

def produce_label(item_list,amount_list,f_id,f_name):
    nutrion_fact={"Formular ID":"",
                  "Formular name":"",
                  "Calories":0,
                  "Total Carbonhydrates":0,
                  "Sugars":0,
                  "Protien":0,
                  "Total Fat":0,
                  "Saturated Fat":0,
                  "Trans Fat":0,
                  "Sodium":0,
                  "Others_list":[],
                  "Others":{},
                  "Serving Size":0
                  }
    df = pd.read_csv("../../../Database/dist/Database/DATA.csv","r",encoding="utf-8",header=None)
    #df = pd.read_csv("./DATA.csv","r",encoding="utf-8",header=None)
    nutrion_fact["Formular ID"]=f_id
    nutrion_fact["Formular name"]=f_name
    for one_item_index in range(len(item_list)):
        for i in range(len(df[0])):
            item_id = df[0][i].split(",")[0]
            if str(item_list[one_item_index]) == item_id:
                nutrion = df[0][i].split(",")[2:10]
                if float(nutrion[0]) != 0:
                    nutrion_fact["Total Carbonhydrates"] +=(float(nutrion[0])/float(100))*float(amount_list[one_item_index])
                    nutrion_fact["Sugars"] +=(float(nutrion[1])/float(100))*float(amount_list[one_item_index])
                    nutrion_fact["Protien"] +=(float(nutrion[2])/float(100))*float(amount_list[one_item_index])
                    nutrion_fact["Total Fat"] +=(float(nutrion[3])/float(100))*float(amount_list[one_item_index])
                    nutrion_fact["Saturated Fat"] +=(float(nutrion[4])/float(100))*float(amount_list[one_item_index])
                    nutrion_fact["Trans Fat"] +=(float(nutrion[5])/float(100))*float(amount_list[one_item_index])
                    nutrion_fact["Sodium"] +=(float(nutrion[6])/float(100))*float(amount_list[one_item_index])
                    nutrion_fact["Serving Size"] +=float(amount_list[one_item_index])
                    if nutrion[7]!="":
                        other_len = len(nutrion[7].split(">"))
                        other_list = nutrion[7].split(">")
                        for x in range(0,other_len,2):
                            nutrion_fact["Others_list"].append(other_list[x])
                            nutrion_fact["Others_list"].append((float(other_list[x+1])*float(amount_list[one_item_index])))
                            
                    print(nutrion_fact["Others_list"])
    nutrion_fact["Total Carbonhydrates"] =round(nutrion_fact["Total Carbonhydrates"],2) 
    nutrion_fact["Sugars"] =round(nutrion_fact["Sugars"],2)   
    nutrion_fact["Protien"] =round(nutrion_fact["Protien"],2)    
    nutrion_fact["Total Fat"] =round(nutrion_fact["Total Fat"],2) 
    nutrion_fact["Saturated Fat"] =round(nutrion_fact["Saturated Fat"],2)
    nutrion_fact["Trans Fat"] =round(nutrion_fact["Trans Fat"],2)
    nutrion_fact["Sodium"] =round(nutrion_fact["Sodium"],2)
    nutrion_fact["Calories"]=round((nutrion_fact["Total Carbonhydrates"]*4+nutrion_fact["Protien"]*4+nutrion_fact["Total Fat"]*9),2)
    #other的是百分比
    n = len(nutrion_fact["Others_list"])
    for x in range(0,n,2):      
        try:
            if nutrion_fact["Others"][nutrion_fact["Others_list"][x]] != 0:
                nutrion_fact["Others"][nutrion_fact["Others_list"][x]]+=(float(nutrion_fact["Others_list"][x+1]))
        except:
            nutrion_fact["Others"][nutrion_fact["Others_list"][x]]=(float(nutrion_fact["Others_list"][x+1]))

#    del nutrion_fact["Others_list"]    
#    nutrion_fact["Others"] =round(nutrion_fact["Others"],2)     
    return nutrion_fact

def produce_excel(label):
    #創建工作簿
    wb = Workbook()
    #獲取活動工作表
    ws = wb.active
    #修改工作表標題
    ws.title = label["Formular ID"]
    
    #欄位標題
    ws.append([])
    ws.append(["","營養標示",label['Formular ID']+label['Formular name']])
    ws.append(["","每一份量",label['Serving Size'],"公克"])
    ws.append(["",' ', '         每份'," ", '         每100公克'," "])
    ws.append(["",'熱量', round(label['Calories'],2),"大卡", round((label['Calories']/label['Serving Size'])*100,2),"大卡"])
    ws.append(["",'蛋白質', round(label['Protien'],2),"公克", round((label['Protien']/label['Serving Size'])*100,2),"公克"])
    ws.append(["",'脂質', round(label['Total Fat'],2),"公克",round((label['Total Fat']/label['Serving Size'])*100,2),"公克"])
    ws.append(["",'    飽和脂肪', round(label['Saturated Fat'],2),"公克",round((label['Saturated Fat']/label['Serving Size'])*100,2),"公克"])
    ws.append(["",'    反式脂肪', round(label['Trans Fat'],2),"公克", round((label['Trans Fat']/label['Serving Size'])*100,2),"公克"])
    ws.append(["",'碳水化合物', round(label['Total Carbonhydrates'],2),"公克", round((label['Total Carbonhydrates']/label['Serving Size'])*100,2),"公克"])
    ws.append(["",'    糖',round(label['Sugars'],2),"公克",round((label['Sugars']/label['Serving Size'])*100,2),"公克"])
    ws.append(["",'鈉', round(label['Sodium']*1000,2),"毫克", round((label['Sodium']*1000/label['Serving Size'])*100,2),"毫克"])
    ws.append(["",'其他'])
    for key in label['Others'].keys():
        ws.append(['',key,round(float(label['Others'][key])*10,2),"毫克",round((float(label['Others'][key])*10/label['Serving Size'])*100,2),"毫克"])
    
    #定義寬高
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 10
    #畫框線
    border_top = Border(top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    top_list = ["C2","C3","C4","D2","D3","D4","E2","E3","E4"]
    for top in top_list: 
        ws[top].border = border_top
    border_top_corner_l = Border(left=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    ws["B2"].border = border_top_corner_l
    ws["B3"].border = border_top_corner_l
    ws["B4"].border = border_top_corner_l
    border_top_corner_r = Border(right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    ws["F2"].border = border_top_corner_r
    ws["F3"].border = border_top_corner_r
    ws["F4"].border = border_top_corner_r
    border_left=Border(left=Side(border_style='thin',color='000000'))
    left_list=["B5","B6","B7","B8","B9","B10","B11","B12"]
    for left in left_list:
        ws[left].border =border_left
    border_right=Border(right=Side(border_style='thin',color='000000'))
    right_list=["F5","F6","F7","F8","F9","F10","F11","F12"]
    for right in right_list:
        ws[right].border =border_right
    border_bottom = Border(bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'))
    bottom_list = ["C13","D13","E13"]
    for bottom in bottom_list: 
        ws[bottom].border = border_bottom
    border_bottom_corner_l = Border(left=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    ws["B13"].border = border_bottom_corner_l
    border_bottom_corner_r = Border(right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
    ws["F13"].border = border_bottom_corner_r
      
     
  
    
    #保存文件
    file_name=label["Formular ID"]
    if label["Formular name"] !="":
        file_name += ("_"+label["Formular name"])
    wb.save("../../../"+file_name+".xlsx")
    #wb.save("./"+file_name+".xlsx")


def fetch(entries):
    item_list = []
    amount_list = []
    fid = f_id.get()
    name = f_name.get()
    for entry in entries:
        item  = entry[1].get()
        amount  = entry[2].get()
        if item != "":
            item_list.append(item)
            amount_list.append(amount)          
        f_id.delete(0,END)
        f_name.delete(0,END)
        entry[1].delete(0,END)
        entry[2].delete(0,END)  
    label = produce_label(item_list,amount_list,fid,name)
    produce_excel(label)
    print(label)
#    df = pd.read_csv("./DATA.csv","r",encoding="utf-8")    
#    for one_item in item_list:
        

ents = makeform(window)

b1 = Button(window, text='Start Calculate!', font=('Arial', 12), width=15, height=1,
          command=(lambda e=ents: fetch(e)))
b1.pack(pady=20)

window.mainloop()






















